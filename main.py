from openpyxl import *
from tkinter import *

wb = load_workbook('C:\\Users\\User\\Desktop\\excel.xlsx')
sheet = wb.active


def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 10

                        ##### definicao do tamanho e nome das celulas

    sheet.cell(row=1, column=1).value = "Nome"
    sheet.cell(row=1, column=2).value = "Endereco"
    sheet.cell(row=1, column=3).value = "Data nascimento"
    sheet.cell(row=1, column=4).value = "Data Batismo"
    sheet.cell(row=1, column=5).value = "Contato"
    sheet.cell(row=1, column=6).value = "Email "
    sheet.cell(row=1, column=7).value = "Estado Civil"



def focus1(event):
    endereco_field.focus_set()


def focus2(event):
    data_nasc_field.focus_set()


def focus3(event):
    data_bat_field.focus_set()


def focus4(event):
    contato_field.focus_set()


def focus5(event):
    email_field.focus_set()


def focus6(event):
    estado_civil_field.focus_set()


def clear():
    nome_field.delete(0, END)
    endereco_field.delete(0, END)
    data_nasc_field.delete(0, END)
    data_bat_field.delete(0, END)
    contato_field.delete(0, END)
    email_field.delete(0, END)
    estado_civil_field.delete(0, END)


def insert():
    if (nome_field.get() == "" and
            endereco_field.get() == "" and
            data_nasc_field.get() == "" and
            data_bat_field.get() == "" and
            contato_field.get() == "" and
            email_field.get() == "" and
            esta_civil_field.get() == ""):
        print("empty input")

    else:

        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = nome_field.get()
        sheet.cell(row=current_row + 1, column=2).value = endereco_field.get()
        sheet.cell(row=current_row + 1, column=3).value = data_nasc_field.get()
        sheet.cell(row=current_row + 1, column=4).value = data_bat_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contato_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_field.get()
        sheet.cell(row=current_row + 1, column=7).value = estado_civil_field.get()
        wb.save('C:\\Users\\User\\Desktop\\excel.xlsx')
        nome_field.focus_set()
        clear()


if __name__ == "__main__":
    root = Tk()

    root.configure(background='light green')

    root.title("Registro Congregação Sul")

    root.geometry("500x300")

    excel()

    heading = Label(root, text="Formulario", bg="light green")

    nome = Label(root, text="Nome", bg="light green")

    endereco = Label(root, text="Endereço", bg="light green")

    data_nasc = Label(root, text="Data Nasc", bg="light green")

    data_bat = Label(root, text="Data Bat", bg="light green")

    contato = Label(root, text="Contato", bg="light green")

    email = Label(root, text="Email", bg="light green")

    estado_civil = Label(root, text="Estado civil", bg="light green")

    heading.grid(row=0, column=1)
    nome.grid(row=1, column=0)
    endereco.grid(row=2, column=0)
    data_nasc.grid(row=3, column=0)
    data_bat.grid(row=4, column=0)
    contato.grid(row=5, column=0)
    email.grid(row=6, column=0)
    estado_civil.grid(row=7, column=0)

    nome_field = Entry(root)
    endereco_field = Entry(root)
    data_nasc_field = Entry(root)
    data_bat_field = Entry(root)
    contato_field = Entry(root)
    email_field = Entry(root)
    estado_civil_field = Entry(root)

    nome_field.bind("<Return>", focus1)

    endereco_field.bind("<Return>", focus2)

    data_nasc_field.bind("<Return>", focus3)

    data_bat_field.bind("<Return>", focus4)

    contato_field.bind("<Return>", focus5)

    email_field.bind("<Return>", focus6)

    nome_field.grid(row=1, column=1, ipadx="100")
    endereco_field.grid(row=2, column=1, ipadx="100")
    data_nasc_field.grid(row=3, column=1, ipadx="100")
    data_bat_field.grid(row=4, column=1, ipadx="100")
    contato_field.grid(row=5, column=1, ipadx="100")
    email_field.grid(row=6, column=1, ipadx="100")
    estado_civil_field.grid(row=7, column=1, ipadx="100")

    excel()

    submit = Button(root, text="Registar", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=8, column=1)

    root.mainloop()